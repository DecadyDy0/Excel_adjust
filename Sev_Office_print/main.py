from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from tempfile import NamedTemporaryFile



app = FastAPI()

@app.get('/inicio')
async def ruta_de_prueba():
    return "Validacion 200 FastAPI"
def eliminar_cuadros_de_texto(archivo_entrada, archivo_salida):
    wb = load_workbook(archivo_entrada)

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        
        for shape in ws._images:
            if 'Made with <3 with APEXOfficePrint (Dev Cred)' in shape.anchor._from.text:
                ws._images.remove(shape)

    wb.save(archivo_salida)

@app.post("/eliminar_cuadros_de_texto/")
async def eliminar_cuadros_de_texto_route(file: UploadFile = File(...)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos Excel (.xlsx)")

    try:
        # Crear un archivo temporal para guardar el archivo de salida
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_output_file:
            temp_output_filename = temp_output_file.name

            # Guardar el archivo subido en disco
            with open(temp_output_filename, "wb") as buffer:
                buffer.write(file.file.read())

            # Llamar a la función para eliminar los cuadros de texto
            eliminar_cuadros_de_texto(temp_output_filename, temp_output_filename)

            # Retornar el archivo ajustado
            return FileResponse(temp_output_filename, filename="documento_ajustado.xlsx")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al procesar el archivo: {str(e)}")
